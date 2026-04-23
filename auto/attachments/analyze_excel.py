import openpyxl
import re

def analyze():
    wb = openpyxl.load_workbook('e:/Chat-Report-Generator-main/esempioRapporto.xlsx', data_only=True)
    sheet = wb.active
    print(f"Sheet Name: {sheet.title}")
    
    header = None
    start_row = 0
    for i, row in enumerate(sheet.iter_rows(max_row=20, values_only=True)):
        row_str = [str(x) for x in row if x]
        if any(re.match(r'Da|A|Corpo|Body|Source|Orientamento', str(x), re.I) for x in row_str):
            header = [str(x) for x in row]
            start_row = i
            print(f"Header found at row {i}")
            break
            
    if not header:
        print("Header NOT found")
        return

    # Print relevant columns
    cols_to_check = ['Da', 'A', 'Corpo', 'Timestamp: Ora', 'Orientamento', 'Stato', 'Tipo', 'Partecipanti']
    for idx, name in enumerate(header):
        if name and any(c.lower() in name.lower() for c in cols_to_check):
            print(f"Column {idx}: {name}")

    # Print first few data rows
    for i, row in enumerate(sheet.iter_rows(min_row=start_row+2, max_row=start_row+10, values_only=True)):
        print(f"Row {i}: {row[:10]}...") # Print first 10 columns
        # specifically look for direction column
        for idx, val in enumerate(row):
            h_name = header[idx] if idx < len(header) else ""
            if h_name and ('Orientamento' in h_name or 'Stato' in h_name or 'Tipo' in h_name or 'Direction' in h_name):
                 print(f"  {h_name} = {val}")

if __name__ == "__main__":
    analyze()
