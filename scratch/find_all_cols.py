import openpyxl

def find_all_cols():
    wb = openpyxl.load_workbook('e:/Chat-Report-Generator-main/esempioRapporto.xlsx', data_only=True)
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"--- Sheet: {name} ---")
        for i, row in enumerate(sheet.iter_rows(max_row=2, values_only=True)):
            print(f"Row {i}: {ascii(row[:30])}") # Peek first 30 columns

if __name__ == "__main__":
    find_all_cols()
